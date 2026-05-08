"""
Excel → SQLite importer for the Project Management schema.

Reads DB_Tables_ProjectManagement.xlsx (one sheet per table) and inserts
every row into the matching SQLite table.

Usage
-----
    python import_excel.py                      # uses data/DB_Tables_ProjectManagement.xlsx
    python import_excel.py /path/to/file.xlsx   # custom location
    python import_excel.py --reset              # wipe all 12 tables first

Notes
-----
- Department.manager_employee_id and Employee.department_id form a cycle.
  We solve it by inserting Departments with manager_employee_id = NULL,
  then inserting Employees, then UPDATEing Departments with their managers.
- We disable FK enforcement during the bulk load and re-enable it at the
  end (after a quick integrity check). This avoids ordering pain inside
  individual sheets while still catching bad references.
- All inserts use the explicit primary key value from the Excel so the FK
  references in later sheets resolve cleanly.
"""

from __future__ import annotations

import os
import sys
from datetime import date, datetime

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required. Install with: pip install openpyxl")

from database import get_connection, init_database, reset_database, DB_PATH

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            'data', 'DB_Tables_ProjectManagement.xlsx')

# Sheet name → (table name, list of columns to insert in order).
# Column order must match the header row of each sheet exactly.
SHEET_MAP = [
    ('Department', 'Department', [
        'department_id', 'department_name', 'manager_employee_id',
        'cost_center', 'status', 'created_at',
    ]),
    ('Employee', 'Employee', [
        'employee_id', 'first_name', 'last_name', 'employee_code',
        'email', 'phone', 'role', 'department_id', 'hourly_rate',
        'hire_date', 'status', 'created_at',
    ]),
    ('External_Resource', 'External_Resource', [
        'external_id', 'first_name', 'last_name', 'company', 'email',
        'phone', 'hourly_rate', 'contract_type', 'specialty', 'status',
        'contract_start', 'contract_end', 'created_at',
    ]),
    ('Customer', 'Customer', [
        'customer_id', 'customer_name', 'customer_code', 'contact_name',
        'contact_email', 'contact_phone', 'industry', 'address', 'country',
        'payment_terms', 'status', 'created_at',
    ]),
    ('Knowledge', 'Knowledge', [
        'knowledge_id', 'subject', 'category', 'description',
    ]),
    ('Project', 'Project', [
        'project_id', 'project_name', 'project_code', 'customer_id',
        'description', 'status', 'priority', 'estimated_hours',
        'actual_hours', 'budget', 'actual_cost', 'start_date', 'end_date',
        'created_at',
    ]),
    ('Task', 'Task', [
        'task_id', 'project_id', 'title', 'description',
        'assigned_employee_id', 'assigned_external_id', 'status',
        'priority', 'estimated_hours', 'actual_hours', 'due_date',
        'completed_date', 'created_at',
    ]),
    ('Project_Assignment', 'Project_Assignment', [
        'assignment_id', 'project_id', 'employee_id', 'external_id',
        'role_in_project', 'allocated_hours', 'start_date', 'end_date',
    ]),
    ('Employee_Knowledge', 'Employee_Knowledge', [
        'employee_knowledge_id', 'employee_id', 'knowledge_id',
        'years_experience', 'proficiency',
    ]),
    ('Time_Entry', 'Time_Entry', [
        'time_entry_id', 'task_id', 'employee_id', 'external_id',
        'date', 'hours', 'description', 'billable',
    ]),
    ('Project_Risk', 'Project_Risk', [
        'risk_id', 'project_id', 'description', 'probability', 'impact',
        'mitigation_plan', 'status', 'identified_date',
    ]),
    ('Invoice', 'Invoice', [
        'invoice_id', 'project_id', 'customer_id', 'invoice_number',
        'invoice_date', 'due_date', 'total_amount', 'status', 'notes',
    ]),
]


def _normalize(value):
    """Convert openpyxl values into types SQLite is happy to store."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        return value.strip() or None
    return value


def _read_sheet(wb, sheet_name, expected_columns):
    """Yield row tuples in `expected_columns` order from a worksheet.

    Validates the header against the expected column list so we fail loudly
    if the Excel evolves underneath us.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    # Truncate to declared length (some sheets may have trailing empties)
    header = header[:len(expected_columns)]
    if header != expected_columns:
        raise ValueError(
            f"Sheet '{sheet_name}' header mismatch.\n"
            f"  expected: {expected_columns}\n"
            f"  got:      {header}"
        )
    for raw in rows[1:]:
        if raw is None or all(v is None for v in raw):
            continue
        # Pad/truncate to column count
        raw = list(raw)[:len(expected_columns)]
        while len(raw) < len(expected_columns):
            raw.append(None)
        yield tuple(_normalize(v) for v in raw)


def import_workbook(xlsx_path: str = DEFAULT_XLSX, *, do_reset: bool = False) -> dict:
    """Load every sheet into its matching SQLite table.

    Returns a dict {table: rows_inserted}.
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    if do_reset:
        reset_database()
    else:
        init_database()

    wb = load_workbook(xlsx_path, data_only=True)
    inserted: dict[str, int] = {}

    with get_connection() as conn:
        cur = conn.cursor()
        # Bulk-load with FKs off; we'll enforce integrity after insert.
        cur.execute("PRAGMA foreign_keys = OFF")

        # --- 1. Department: insert WITHOUT manager_employee_id (cycle break) ---
        sheet, table, cols = SHEET_MAP[0]
        manager_updates = []  # (department_id, manager_employee_id)
        dept_cols = [c for c in cols if c != 'manager_employee_id']
        placeholders = ','.join('?' for _ in dept_cols)
        sql = f"INSERT OR REPLACE INTO {table} ({','.join(dept_cols)}) VALUES ({placeholders})"
        n = 0
        for row in _read_sheet(wb, sheet, cols):
            row_dict = dict(zip(cols, row))
            if row_dict['manager_employee_id'] is not None:
                manager_updates.append(
                    (row_dict['department_id'], row_dict['manager_employee_id'])
                )
            cur.execute(sql, tuple(row_dict[c] for c in dept_cols))
            n += 1
        inserted[table] = n

        # --- 2..N: every other sheet, straightforward ---
        for sheet, table, cols in SHEET_MAP[1:]:
            placeholders = ','.join('?' for _ in cols)
            sql = f"INSERT OR REPLACE INTO {table} ({','.join(cols)}) VALUES ({placeholders})"
            n = 0
            for row in _read_sheet(wb, sheet, cols):
                cur.execute(sql, row)
                n += 1
            inserted[table] = n

        # --- 3. Resolve the cycle: set manager_employee_id on Department ---
        cur.executemany(
            "UPDATE Department SET manager_employee_id = ? WHERE department_id = ?",
            [(mgr, dept) for dept, mgr in manager_updates],
        )

        # --- 4. Re-enable FKs and check integrity ---
        cur.execute("PRAGMA foreign_keys = ON")
        cur.execute("PRAGMA foreign_key_check")
        violations = cur.fetchall()
        if violations:
            raise RuntimeError(
                f"FK integrity check failed after import: {len(violations)} violations. "
                f"First few: {[dict(v) for v in violations[:5]]}"
            )

    return inserted


def main(argv):
    do_reset = '--reset' in argv
    args = [a for a in argv if not a.startswith('--')]
    xlsx = args[0] if args else DEFAULT_XLSX

    print(f"→ Excel : {xlsx}")
    print(f"→ DB    : {DB_PATH}")
    print(f"→ Reset : {do_reset}")
    print()

    inserted = import_workbook(xlsx, do_reset=do_reset)
    total = sum(inserted.values())
    print(f"✓ Imported {total} rows across {len(inserted)} tables:")
    for table, n in inserted.items():
        print(f"  {table:<22} {n:>4} rows")


if __name__ == '__main__':
    main(sys.argv[1:])
