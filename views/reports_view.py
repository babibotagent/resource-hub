"""
Reports View - Relatórios e exportação
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import database as db


class ReportsView:
    def __init__(self, parent, colors):
        self.parent = parent
        self.colors = colors
        self.frame = tk.Frame(parent, bg=colors['background'])
        self.frame.pack(fill='both', expand=True)
        self._build()
    
    def _build(self):
        # Header
        header = tk.Frame(self.frame, bg=self.colors['background'])
        header.pack(fill='x', padx=30, pady=(25, 15))
        
        title_frame = tk.Frame(header, bg=self.colors['background'])
        title_frame.pack(side='left')
        
        tk.Label(title_frame, text="Relatórios",
                font=('Segoe UI', 24, 'bold'),
                bg=self.colors['background'],
                fg=self.colors['text']).pack(anchor='w')
        
        tk.Label(title_frame, text="Gere e exporte relatórios em Excel",
                font=('Segoe UI', 10),
                bg=self.colors['background'],
                fg=self.colors['text_light']).pack(anchor='w')
        
        # Container principal
        main_container = tk.Frame(self.frame, bg=self.colors['background'])
        main_container.pack(fill='both', expand=True, padx=30, pady=(0, 30))
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=1)
        
        # Cards de relatórios
        reports = [
            {
                'icon': '📁',
                'title': 'Relatório de Projetos',
                'description': 'Lista completa de projetos com status, orçamento, datas e gerentes',
                'action': self._export_projects,
                'color': self.colors['primary']
            },
            {
                'icon': '👥',
                'title': 'Relatório de Recursos',
                'description': 'Cadastro completo de recursos com função, taxa horária e capacidade',
                'action': self._export_resources,
                'color': self.colors['success']
            },
            {
                'icon': '🔗',
                'title': 'Relatório de Alocações',
                'description': 'Todas as alocações com horas planejadas vs realizadas',
                'action': self._export_allocations,
                'color': self.colors['warning']
            },
            {
                'icon': '📊',
                'title': 'Relatório Consolidado',
                'description': 'Dashboard executivo com KPIs e múltiplas abas',
                'action': self._export_consolidated,
                'color': self.colors['danger']
            },
            {
                'icon': '⏱️',
                'title': 'Utilização de Recursos',
                'description': 'Análise de capacidade e utilização por recurso',
                'action': self._export_utilization,
                'color': self.colors['secondary']
            },
            {
                'icon': '💰',
                'title': 'Análise de Custos',
                'description': 'Custos planejados vs realizados por projeto',
                'action': self._export_costs,
                'color': '#8b5cf6'
            },
        ]
        
        for i, report in enumerate(reports):
            row = i // 2
            col = i % 2
            self._create_report_card(main_container, report, row, col)
    
    def _create_report_card(self, parent, report, row, col):
        """Cria um card de relatório"""
        card = tk.Frame(parent, bg='white', highlightbackground=self.colors['border'],
                       highlightthickness=1)
        card.grid(row=row, column=col, sticky='nsew', padx=10, pady=10)
        
        # Barra colorida superior
        tk.Frame(card, bg=report['color'], height=4).pack(fill='x')
        
        # Conteúdo
        content = tk.Frame(card, bg='white')
        content.pack(fill='both', expand=True, padx=25, pady=20)
        
        # Ícone
        tk.Label(content, text=report['icon'],
                font=('Segoe UI', 32),
                bg='white', fg=report['color']).pack(anchor='w')
        
        # Título
        tk.Label(content, text=report['title'],
                font=('Segoe UI', 14, 'bold'),
                bg='white', fg=self.colors['text']).pack(anchor='w', pady=(10, 5))
        
        # Descrição
        tk.Label(content, text=report['description'],
                font=('Segoe UI', 9),
                bg='white', fg=self.colors['text_light'],
                wraplength=400, justify='left').pack(anchor='w', pady=(0, 15))
        
        # Botão
        tk.Button(content, text="📥 Exportar Excel",
                 font=('Segoe UI', 10, 'bold'),
                 bg=report['color'], fg='white',
                 bd=0, padx=20, pady=10, cursor='hand2',
                 command=report['action']).pack(anchor='w')
    
    def _ask_save_path(self, default_name):
        """Solicita caminho para salvar"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            initialfile=f"{default_name}_{timestamp}.xlsx"
        )
    
    def _export_projects(self):
        """Exporta relatório de projetos"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            path = self._ask_save_path('Relatorio_Projetos')
            if not path:
                return
            
            projects = db.get_all_projects()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Projetos"
            
            # Cabeçalhos
            headers = ['ID', 'Nome', 'Cliente', 'Gerente', 'Início', 'Fim',
                      'Orçamento (R$)', 'Status', 'Prioridade', 'Descrição']
            
            self._apply_header_style(ws, headers)
            
            # Dados
            for row_idx, p in enumerate(projects, start=2):
                ws.cell(row=row_idx, column=1, value=p['id'])
                ws.cell(row=row_idx, column=2, value=p['name'])
                ws.cell(row=row_idx, column=3, value=p['client'] or '-')
                ws.cell(row=row_idx, column=4, value=p['manager'] or '-')
                ws.cell(row=row_idx, column=5, value=p['start_date'] or '-')
                ws.cell(row=row_idx, column=6, value=p['end_date'] or '-')
                ws.cell(row=row_idx, column=7, value=p['budget'] or 0)
                ws.cell(row=row_idx, column=8, value=p['status'])
                ws.cell(row=row_idx, column=9, value=p['priority'])
                ws.cell(row=row_idx, column=10, value=p['description'] or '-')
            
            self._auto_size_columns(ws)
            wb.save(path)
            
            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso!\n\n{path}")
            self._open_file(path)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
    
    def _export_resources(self):
        """Exporta relatório de recursos"""
        try:
            from openpyxl import Workbook
            
            path = self._ask_save_path('Relatorio_Recursos')
            if not path:
                return
            
            resources = db.get_all_resources()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Recursos"
            
            headers = ['ID', 'Nome', 'Função', 'Tipo', 'Email', 'Taxa Horária (R$)',
                      'Capacidade Semanal (h)', 'Habilidades', 'Status']
            
            self._apply_header_style(ws, headers)
            
            for row_idx, r in enumerate(resources, start=2):
                ws.cell(row=row_idx, column=1, value=r['id'])
                ws.cell(row=row_idx, column=2, value=r['name'])
                ws.cell(row=row_idx, column=3, value=r['role'] or '-')
                ws.cell(row=row_idx, column=4, value=r['resource_type'])
                ws.cell(row=row_idx, column=5, value=r['email'] or '-')
                ws.cell(row=row_idx, column=6, value=r['hourly_rate'] or 0)
                ws.cell(row=row_idx, column=7, value=r['capacity_hours_week'] or 0)
                ws.cell(row=row_idx, column=8, value=r['skills'] or '-')
                ws.cell(row=row_idx, column=9, value=r['status'])
            
            self._auto_size_columns(ws)
            wb.save(path)
            
            messagebox.showinfo("Sucesso", f"Relatório exportado!\n\n{path}")
            self._open_file(path)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
    
    def _export_allocations(self):
        """Exporta relatório de alocações"""
        try:
            from openpyxl import Workbook
            
            path = self._ask_save_path('Relatorio_Alocacoes')
            if not path:
                return
            
            allocations = db.get_all_allocations()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Alocações"
            
            headers = ['ID', 'Projeto', 'Recurso', 'Função', 'Início', 'Fim',
                      'Horas Planejadas', 'Horas Realizadas', 'Progresso (%)',
                      'Variação (h)', 'Observações']
            
            self._apply_header_style(ws, headers)
            
            for row_idx, a in enumerate(allocations, start=2):
                planned = a['planned_hours'] or 0
                actual = a['actual_hours'] or 0
                progress = (actual / planned * 100) if planned > 0 else 0
                variation = actual - planned
                
                ws.cell(row=row_idx, column=1, value=a['id'])
                ws.cell(row=row_idx, column=2, value=a['project_name'])
                ws.cell(row=row_idx, column=3, value=a['resource_name'])
                ws.cell(row=row_idx, column=4, value=a['role_in_project'] or '-')
                ws.cell(row=row_idx, column=5, value=a['start_date'])
                ws.cell(row=row_idx, column=6, value=a['end_date'])
                ws.cell(row=row_idx, column=7, value=planned)
                ws.cell(row=row_idx, column=8, value=actual)
                ws.cell(row=row_idx, column=9, value=round(progress, 2))
                ws.cell(row=row_idx, column=10, value=variation)
                ws.cell(row=row_idx, column=11, value=a['notes'] or '-')
            
            self._auto_size_columns(ws)
            wb.save(path)
            
            messagebox.showinfo("Sucesso", f"Relatório exportado!\n\n{path}")
            self._open_file(path)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
    
    def _export_consolidated(self):
        """Exporta relatório consolidado em múltiplas abas"""
        try:
            from openpyxl import Workbook
            
            path = self._ask_save_path('Relatorio_Consolidado')
            if not path:
                return
            
            wb = Workbook()
            
            # Aba 1: Resumo Executivo
            ws_summary = wb.active
            ws_summary.title = "Resumo Executivo"
            
            stats = db.get_dashboard_stats()
            
            ws_summary['A1'] = "RELATÓRIO CONSOLIDADO - GESTÃO DE RECURSOS"
            ws_summary['A1'].font = self._title_font()
            ws_summary.merge_cells('A1:D1')
            
            ws_summary['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws_summary.merge_cells('A2:D2')
            
            ws_summary['A4'] = "INDICADORES PRINCIPAIS"
            ws_summary['A4'].font = self._subtitle_font()
            
            indicators = [
                ('Total de Projetos', stats['total_projects']),
                ('Projetos Ativos', stats['active_projects']),
                ('Total de Recursos Ativos', stats['total_resources']),
                ('Horas Planejadas', f"{stats['planned_hours']:.0f}h"),
                ('Horas Realizadas', f"{stats['actual_hours']:.0f}h"),
                ('Orçamento Total', f"R$ {stats['total_budget']:,.2f}"),
                ('Custo Realizado', f"R$ {stats['actual_cost']:,.2f}"),
            ]
            
            for i, (label, value) in enumerate(indicators, start=5):
                ws_summary.cell(row=i, column=1, value=label).font = self._bold_font()
                ws_summary.cell(row=i, column=2, value=value)
            
            # Aba 2: Projetos
            ws_p = wb.create_sheet("Projetos")
            headers = ['ID', 'Nome', 'Cliente', 'Gerente', 'Início', 'Fim',
                      'Orçamento', 'Status', 'Prioridade']
            self._apply_header_style(ws_p, headers)
            
            for row_idx, p in enumerate(db.get_all_projects(), start=2):
                ws_p.cell(row=row_idx, column=1, value=p['id'])
                ws_p.cell(row=row_idx, column=2, value=p['name'])
                ws_p.cell(row=row_idx, column=3, value=p['client'] or '-')
                ws_p.cell(row=row_idx, column=4, value=p['manager'] or '-')
                ws_p.cell(row=row_idx, column=5, value=p['start_date'])
                ws_p.cell(row=row_idx, column=6, value=p['end_date'])
                ws_p.cell(row=row_idx, column=7, value=p['budget'] or 0)
                ws_p.cell(row=row_idx, column=8, value=p['status'])
                ws_p.cell(row=row_idx, column=9, value=p['priority'])
            
            self._auto_size_columns(ws_p)
            
            # Aba 3: Recursos
            ws_r = wb.create_sheet("Recursos")
            headers = ['ID', 'Nome', 'Função', 'Tipo', 'Taxa Horária', 'Capacidade', 'Status']
            self._apply_header_style(ws_r, headers)
            
            for row_idx, r in enumerate(db.get_all_resources(), start=2):
                ws_r.cell(row=row_idx, column=1, value=r['id'])
                ws_r.cell(row=row_idx, column=2, value=r['name'])
                ws_r.cell(row=row_idx, column=3, value=r['role'] or '-')
                ws_r.cell(row=row_idx, column=4, value=r['resource_type'])
                ws_r.cell(row=row_idx, column=5, value=r['hourly_rate'] or 0)
                ws_r.cell(row=row_idx, column=6, value=r['capacity_hours_week'] or 0)
                ws_r.cell(row=row_idx, column=7, value=r['status'])
            
            self._auto_size_columns(ws_r)
            
            # Aba 4: Alocações
            ws_a = wb.create_sheet("Alocações")
            headers = ['ID', 'Projeto', 'Recurso', 'Início', 'Fim',
                      'Horas Plan.', 'Horas Real.', 'Progresso (%)']
            self._apply_header_style(ws_a, headers)
            
            for row_idx, a in enumerate(db.get_all_allocations(), start=2):
                planned = a['planned_hours'] or 0
                actual = a['actual_hours'] or 0
                progress = (actual / planned * 100) if planned > 0 else 0
                
                ws_a.cell(row=row_idx, column=1, value=a['id'])
                ws_a.cell(row=row_idx, column=2, value=a['project_name'])
                ws_a.cell(row=row_idx, column=3, value=a['resource_name'])
                ws_a.cell(row=row_idx, column=4, value=a['start_date'])
                ws_a.cell(row=row_idx, column=5, value=a['end_date'])
                ws_a.cell(row=row_idx, column=6, value=planned)
                ws_a.cell(row=row_idx, column=7, value=actual)
                ws_a.cell(row=row_idx, column=8, value=round(progress, 2))
            
            self._auto_size_columns(ws_a)
            
            wb.save(path)
            
            messagebox.showinfo("Sucesso", f"Relatório consolidado exportado!\n\n{path}")
            self._open_file(path)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
    
    def _export_utilization(self):
        """Exporta análise de utilização"""
        try:
            from openpyxl import Workbook
            
            path = self._ask_save_path('Utilizacao_Recursos')
            if not path:
                return
            
            utilization = db.get_resource_utilization()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Utilização"
            
            headers = ['ID', 'Recurso', 'Capacidade Semanal (h)',
                      'Horas Planejadas', 'Horas Realizadas',
                      'Utilização (%)', 'Eficiência (%)']
            
            self._apply_header_style(ws, headers)
            
            for row_idx, u in enumerate(utilization, start=2):
                planned = u['planned_hours'] or 0
                actual = u['actual_hours'] or 0
                utilization_pct = (actual / planned * 100) if planned > 0 else 0
                efficiency = (planned / actual * 100) if actual > 0 else 0
                
                ws.cell(row=row_idx, column=1, value=u['id'])
                ws.cell(row=row_idx, column=2, value=u['name'])
                ws.cell(row=row_idx, column=3, value=u['capacity_hours_week'])
                ws.cell(row=row_idx, column=4, value=planned)
                ws.cell(row=row_idx, column=5, value=actual)
                ws.cell(row=row_idx, column=6, value=round(utilization_pct, 2))
                ws.cell(row=row_idx, column=7, value=round(efficiency, 2))
            
            self._auto_size_columns(ws)
            wb.save(path)
            
            messagebox.showinfo("Sucesso", f"Relatório exportado!\n\n{path}")
            self._open_file(path)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
    
    def _export_costs(self):
        """Exporta análise de custos"""
        try:
            from openpyxl import Workbook
            
            path = self._ask_save_path('Analise_Custos')
            if not path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Análise de Custos"
            
            headers = ['Projeto', 'Orçamento', 'Custo Planejado',
                      'Custo Realizado', 'Variação (R$)', 'Variação (%)', 'Status']
            
            self._apply_header_style(ws, headers)
            
            projects = db.get_all_projects()
            row_idx = 2
            
            for project in projects:
                allocations = db.get_allocations_by_project(project['id'])
                
                planned_cost = 0
                actual_cost = 0
                
                for alloc in allocations:
                    resource = db.get_resource(alloc['resource_id'])
                    if resource:
                        rate = resource['hourly_rate'] or 0
                        planned_cost += (alloc['planned_hours'] or 0) * rate
                        actual_cost += (alloc['actual_hours'] or 0) * rate
                
                variation = actual_cost - planned_cost
                variation_pct = (variation / planned_cost * 100) if planned_cost > 0 else 0
                
                ws.cell(row=row_idx, column=1, value=project['name'])
                ws.cell(row=row_idx, column=2, value=project['budget'] or 0)
                ws.cell(row=row_idx, column=3, value=round(planned_cost, 2))
                ws.cell(row=row_idx, column=4, value=round(actual_cost, 2))
                ws.cell(row=row_idx, column=5, value=round(variation, 2))
                ws.cell(row=row_idx, column=6, value=round(variation_pct, 2))
                ws.cell(row=row_idx, column=7, value=project['status'])
                
                row_idx += 1
            
            self._auto_size_columns(ws)
            wb.save(path)
            
            messagebox.showinfo("Sucesso", f"Relatório exportado!\n\n{path}")
            self._open_file(path)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
    
    # ============== HELPERS ==============
    def _apply_header_style(self, ws, headers):
        """Aplica estilo aos cabeçalhos"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
    
    def _auto_size_columns(self, ws):
        """Ajusta automaticamente o tamanho das colunas"""
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
            except AttributeError:
                # Para células mescladas
                continue
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
    
    def _title_font(self):
        from openpyxl.styles import Font
        return Font(name='Calibri', size=16, bold=True, color='2563EB')
    
    def _subtitle_font(self):
        from openpyxl.styles import Font
        return Font(name='Calibri', size=12, bold=True, color='1E40AF')
    
    def _bold_font(self):
        from openpyxl.styles import Font
        return Font(name='Calibri', size=11, bold=True)
    
    def _open_file(self, path):
        """Tenta abrir o arquivo gerado"""
        try:
            import platform
            import subprocess
            system = platform.system()
            if system == 'Windows':
                os.startfile(path)
            elif system == 'Darwin':
                subprocess.call(['open', path])
            else:
                subprocess.call(['xdg-open', path])
        except Exception:
            pass  # Se não conseguir abrir, ignora
